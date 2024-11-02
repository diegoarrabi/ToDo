from config import getPaths, clearFolder, tableStyle, excelPath
import csv
import os
import sys
import openpyxl
from subprocess import call
import pandas as pd
import datetime
import dataframe_image as dfi

def main():
    pDict = getPaths()
    tableIMGdir = pDict['tableIMG']
    projectDir = pDict['Project']
    
    clearFolder(tableIMGdir)

    workbookName = excelPath()
    sheetName = "Assignments"
    wb = openExcelSheet(workbookName, sheetName)
    startTable = findMatrixLU(wb)
    assignmentList = makeAssg(wb, startTable)
    headerCol = getHeader(assignmentList)
    twoWeekList = fullFormat(assignmentList)
    saveName = timeLabel()

    df = pd.DataFrame(twoWeekList, columns=[str(headerCol[0]), str(headerCol[1]), str(headerCol[2])])
    tableDone = df.style.set_table_styles(styleTable(df, headerCol)).hide()
    dfi.export(tableDone, f'{tableIMGdir}/{saveName}.png', dpi=300)

    pyName = "makeWallpaper.py"
    script = os.path.join(projectDir, pyName)
    call(['python3', script])
    return

def timeLabel():
    def getTime():
        tdHour = datetime.datetime.today().hour
        tdMin = datetime.datetime.today().minute
        tdSec = datetime.datetime.today().second
        timeList = [tdHour, tdMin, tdSec]
        return padNum(timeList)
            
    def padNum(var):
        for i in range(len(var)):
            timeVar = str(var[i])
            if len(timeVar) < 2:
                timeVar = "0" + timeVar
            var[i] = timeVar
        return comboTime(var)

    def comboTime(var):
        timeStr = ""
        for i in range(len(var)):
            timeStr = timeStr + var[i]
        return timeStr
    return getTime()

def errorOccured(errorVar, wbName="", basedir="", excelPath=""):
    if errorVar == 1:
        print("----File Not Found----")
        print("-----------------------")
        print(f"File: {wbName}")
        print(f"Directory: {basedir}")
        print("-----------------------")
        print(f"FullPath: {excelPath}")
        print("-----------------------")
    elif errorVar == 2:
        print(f"Sheet: '{wbName}' not found")
    sys.exit()
    return

def openExcelSheet(wbName, ssName):
    workbook = openpyxl.load_workbook(wbName, data_only=True)
    sheetInt = isofInterest(workbook, ssName)
    """wbName = "/" + str(wbName)
    basedir = os.path.dirname(__file__)
    excelPath = os.path.join(basedir+wbName)
    try:
        workbook = openpyxl.load_workbook(excelPath, data_only=True)
    except FileNotFoundError:
        errorOccured(1, wbName, basedir, excelPath)
        return
    sheetInt = isofInterest(workbook, ssName)"""
    return sheetInt

def isofInterest(wb, ssName):
    for sheets in wb:
        sName = str(sheets)
        sName = sName.split('"', 2)
        if sName[1] == str(ssName):
            wbSheet = sheets
            return wbSheet
    errorOccured(2, ssName)

def findMatrixLU(wb):
    def checkColumn(wb, numVar):
        columnValue = True
        i = 0
        while columnValue != None:
            cellColumn = wb.cell(numVar, numVar-i).value
            if cellColumn == None:
                colIndex = (numVar-i+1)
                return colIndex
            i += 1
            
    def checkRow(wb, numVar):
        rowValue = True
        i = 0
        while rowValue != None:
            cellRow = wb.cell(numVar-i, numVar).value
            if cellRow == None:
                rowIndex = (numVar-i+1)
                return rowIndex
            i += 1

    def assigCount(wb, rowIndex, colIndex):
        rowStart = rowIndex
        rowValue = False
        i = 0
        while rowValue == False:
            cellRow = wb.cell(rowIndex+i, colIndex+3).value
            if cellRow == None:
                if wb.cell(rowIndex+i+1, colIndex+3).value != None:
                    pass
                elif wb.cell(rowIndex+i+1, colIndex+3).value == None:
                    return (rowIndex+i-rowStart)
            i += 1
        return

    hasValue = False
    i = 1
    rowI = 0
    while hasValue == False:
        cellMatrix = wb.cell(rowI + i,i)
        cellValue = cellMatrix.value
        if i > 20:
            rowI += 1
            i = 0
        if cellValue != None:
            colIndex = checkColumn(wb, cellMatrix.column) 
            rowIndex = checkRow(wb, cellMatrix.row)
            lastCell = assigCount(wb, rowIndex, colIndex)
            hasValue = True
        i += 1 
    return rowIndex, colIndex, lastCell

def makeAssg(wb, startTable):
    assignmentList = []
    rowInd = startTable[0]
    colInd = startTable[1]
    assLen = startTable[2]
    for i in range(assLen):
        indAssignment = []
        for l in range(5):
            cellIndiv = wb.cell(rowInd+i,colInd+l).value
            indAssignment.append(cellIndiv)
        assignmentList.append(indAssignment)
    return assignmentList

def getHeader(assignmentList):
        newHeader = []
        assignmentHeader = assignmentList.pop(0)
        newHeader.append(assignmentHeader[3])
        newHeader.append(assignmentHeader[2])
        newHeader.append(assignmentHeader[1])
        return newHeader

def fullFormat(assignmentList):

    def formatList(assignmentList):
        removeBlank = [item for item in assignmentList if item[3] != None]
        removeDone = [item for item in removeBlank if item[4] != "done"]
        return removeDone
    
    def listDays(listToSort):
        return listToSort[1]
    
    def formatTime(dateVar):
        def padNum(var):
            for i in range(len(var)):
                timeVar = str(var[i])
                if len(timeVar) < 2:
                    timeVar = "0" + timeVar
                var[i] = timeVar
            return comboTime(var)
        def comboTime(var):
            timeStr = "/"
            dateStr = var[0] + timeStr + var[1]
            return dateStr

        tdMonth = dateVar.month
        tdDay = dateVar.day
        dateList = [tdMonth, tdDay]
        return padNum(dateList)
    
    def make2Week():
        newMasterAssig = []
        for assignInfo in formatMasterList:    
            if assignInfo[1] < 8:
                newAssig = []
                newAssig.append(assignInfo[3])
                newAssig.append(assignInfo[2])
                newAssig.append(assignInfo[1])
                newMasterAssig.append(newAssig)
        return newMasterAssig
    
    def getDaysBW(twoWeekList):
        todayDate = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        for i in twoWeekList:
            someDate = i[1]
            timeInBw = (str(someDate - todayDate))
            if timeInBw.startswith('-'):
                timeBW = timeInBw.split(' ')
                if int(timeBW[0]) == -1:
                    i[2] = 'YESTERDAY!'
                    i[0] = str(i[0]).upper()
                else:
                    i[2] = f'{timeBW[0]} DAYS AGO!'
                    i[0] = str(i[0]).upper()            
            elif timeInBw.startswith('0:0'):
                i[2] = "TODAY!!!"
                i[0] = str(i[0]).upper()
            elif timeInBw.startswith('1 '):
                i[2] = "Tomorrow"
            else:
                timeBW = timeInBw.split(',')
                i[2] = timeBW[0]
        return twoWeekList
    
    def convertDate(twoWeekList):
        for i in twoWeekList:   
            newDate = formatTime(i[1])
            i[1] = newDate
        return twoWeekList 
        
    formatMasterList = formatList(assignmentList)
    formatMasterList.sort(key=listDays)
    twoWeek = getDaysBW(make2Week())
    finalList = convertDate(twoWeek)
    return finalList

def styleTable(df, headerCol):
    cStyle = tableStyle()
    bPx = cStyle['brWidth']
    bCo = cStyle['brColor']
    hFnt = cStyle['fontHead']
    hSFnt = cStyle['fontHsize']
    bFnt = cStyle['fontBody']
    bSFnt = cStyle['fontBsize']
    hCo = cStyle['HeadColor']
    hFntCo = cStyle['hFntColor']
    bFntCo = cStyle['bFntColor']
    rECo = cStyle['rowCoE']
    rOCo = cStyle['rowCoO']

    paddingHead = 'padding-top: %sem; padding-bottom: %sem;' % (0, 0)
    propsHead = 'font-family: %s; color: #%s; font-size: %sem;' % (hFnt, hFntCo, hSFnt)
    
    paddingBody = 'padding-top: %sem; padding-bottom: %sem;' % (0.3, 0.3)
    paddingBodyL = 'padding-left: %sem; padding-top: %sem; padding-bottom: %sem;' % (1, 0.3, 0.3)
    propsBodyE = 'background-color: #%s; font-family: %s; color: #%s; font-size: %sem;'% (rECo, bFnt, bFntCo, bSFnt)
    propsBodyO = 'background-color: #%s; font-family: %s; color: #%s; font-size: %sem;'% (rOCo, bFnt, bFntCo, bSFnt)

    bhBottom = 'border-bottom: %spx solid #%s;' % ((bPx-2), '272727')
    bTop = 'border-top: %spx solid #%s;' % (bPx, bCo)
    bRight = 'border-right: %spx solid #%s;'% (bPx, bCo)
    bBottom = 'border-bottom: %spx solid #%s;'% (bPx, bCo)
    bLeft = 'border-left: %spx solid #%s;'% (bPx, bCo)

    styleList = [
    ##COLOR
        {"selector":"th.col_heading",
            "props": f'background-color: #{hCo}; {propsHead}; {paddingHead}; {bTop}; {bhBottom};'},
            
        {"selector":"tbody tr:nth-child(even)",
            "props": f'{propsBodyE};'},

        {"selector":"tbody tr:nth-child(odd)",
         "props": f'{propsBodyO};'},
    ##ALIGNMENT
        {"selector":"th.col0",
            "props": f"text-align: left;{bLeft}; {paddingBodyL};"},
        
        {"selector":"td.col0",
            "props": f"text-align: left; {paddingBodyL}; {bLeft};"},

        {"selector":"th.col1",
            "props": "text-align: center;"},
        
        {"selector":"td.col1",
            "props": f"text-align: center; {paddingBody}"},

        {"selector":"th.col2",
            "props": f"text-align: center; {bRight}"},

        {"selector":"td.col2",
            "props": f"text-align: right; {paddingBody}; {bRight}"},
        
        {"selector":"tbody tr:nth-last-child(1)", 
            "props": f"text-align: right; {paddingBody}; {bBottom}"},
            ]
    
    lengthofToday = len(df[df[headerCol[2]].str.contains("!")])
    if lengthofToday > 0:
        for i in range(lengthofToday):
            tempToday = {}
            dict_keys = ['selector', 'props']
            tempToday[dict_keys[0]] = ""
            tempToday[dict_keys[1]] = 'font-weight: bold; color: #%s;' % (cStyle['pastCo'])
            tempToday['selector'] = (f"tbody tr:nth-child({i+1})")
            styleList.append(tempToday)
            del tempToday
    return styleList

#--------------------------------------------------------------#

if __name__ == '__main__':
    main()


"""
basedir = os.path.dirname(__file__)
savePath = basedir + "/ToDo_Table/"
script = (savePath)

folderFiles = os.listdir(savePath)

for i in folderFiles:
    remPath = savePath + i
    os.remove(remPath)

workbookName = "pyToDo.xlsm"
sheetName = "Assignments"
wb = openExcelSheet(workbookName, sheetName)
startTable = findMatrixLU(wb)
assignmentList = makeAssg(wb, startTable)
headerCol = getHeader(assignmentList)
twoWeekList = fullFormat(assignmentList)
saveName = timeLabel()

df = pd.DataFrame(twoWeekList, columns=[str(headerCol[0]), str(headerCol[1]), str(headerCol[2])])
tableDone = df.style.set_table_styles(styleTable(df, headerCol)).hide()
dfi.export(tableDone, f'ToDo_Table/{saveName}.png', dpi=300)



script = ("/Users/diegoibarra/Coding/1_Python/Projects/0_PlayProjects/ExcelToDo/saveWallpaper.py")
subprocess.call(['python3', script])
"""