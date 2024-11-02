from config import *
import os
from os import path
import sys
import openpyxl
import datetime
import pandas as pd
from subprocess import run
import dataframe_image as dfi


def main():

    project_dict = getPaths()
    projectDir = project_dict['Project']
    images_dir = project_dict['images']

    deletePreviousTable(images_dir)

    workbookName = excelPath(projectDir)
    sheetname = "Assignments"

    open_workbook = openExcelSheet(workbookName, sheetname)

    table_indices = findMatrixLU(open_workbook)
    assignmentList = makeAssg(open_workbook, table_indices)
    header = getHeader(assignmentList)
    twoWeekList = fullFormat(assignmentList, day_limit)
    
    saveName = 'table'

    """
    print(f'\nStart table: {table_indices}\n')
    print(f'Assignment List: {assignmentList}\n')
    print(f'Header Row: {headerRow}\n')
    print(f'TwoWeekList: {twoWeekList}\n')
    print(f'SaveName: {saveName}\n')
    """

    dataframe = pd.DataFrame(twoWeekList, columns=[str(
        header[0]), str(header[1]), str(header[2])])
    table_finished = dataframe.style.set_table_styles(
        styleTable(dataframe, header)).hide()

    if len(twoWeekList) != 0:
        dfi.export(table_finished, f'{images_dir}/{saveName}.png', dpi=300)

    python_script = "makeWallpaper.py"
    script = os.path.join(projectDir, python_script)

    run(['python3', script])
    return

# --------------------------------------------------------------------------------------------


def deletePreviousTable(images_dir: str) -> None:
    previous_table = [x for x in listdir(images_dir) if x.startswith("table")]
    if len(previous_table) == 0:
        return
    previous_path = path.join(images_dir, previous_table[0])
    run(["rm", "-f", previous_path])


def timeLabel():
    def getTime():
        tdHour = datetime.datetime.today().hour
        tdMin = datetime.datetime.today().minute
        tdSec = datetime.datetime.today().second
        timeList = [tdHour, tdMin, tdSec]
        return padNum(timeList)

    def padNum(var):
        for i in range(len(var)):
            timeVar = str(var[i])
            if len(timeVar) < 2:
                timeVar = "0" + timeVar
            var[i] = timeVar
        return comboTime(var)

    def comboTime(var):
        timeStr = ""
        for i in range(len(var)):
            timeStr = timeStr + var[i]
        return timeStr
    return getTime()


def errorOccured(errorVar, wbName="", basedir="", excelPath=""):
    if errorVar == 1:
        print("----File Not Found----")
        print("-----------------------")
        print(f"File: {wbName}")
        print(f"Directory: {basedir}")
        print("-----------------------")
        print(f"FullPath: {excelPath}")
        print("-----------------------")
    elif errorVar == 2:
        print(f"Sheet: '{wbName}' not found")
    sys.exit()
    return


def openExcelSheet(wbName, ssName):
    workbook = openpyxl.load_workbook(wbName)
    workbook.save(wbName)
    workbook = openpyxl.load_workbook(wbName, data_only=True)
    sheetInt = isofInterest(workbook, ssName)
    return sheetInt


def isofInterest(wb, ssName):
    for sheets in wb:
        sName = str(sheets)
        sName = sName.split('"', 2)
        if sName[1] == str(ssName):
            wbSheet = sheets
            return wbSheet
    errorOccured(2, ssName)


def findMatrixLU(wb):

    def findMatrixMain(wb):

        column_index = 1
        row_index = 1
        last_cell = assigCount(wb, row_index, column_index)

        return row_index, column_index, last_cell

    def assigCount(wb, rowIndex, colIndex):
        rowStart = rowIndex
        rowValue = False
        i = 0
        while rowValue == False:
            cellRow = wb.cell(rowIndex+i, colIndex+1).value
            if cellRow == None:
                if wb.cell(rowIndex+i+1, colIndex+1).value != None:
                    pass
                elif wb.cell(rowIndex+i+1, colIndex+1).value == None:
                    return (rowIndex+i-rowStart)
            i += 1
        return

    return findMatrixMain(wb)


def makeAssg(wb, startTable):
    assignmentList = []
    rowInd = startTable[0]
    colInd = startTable[1]
    assLen = startTable[2]
    for i in range(assLen):
        indAssignment = []
        for l in range(3):
            cellIndiv = wb.cell(rowInd+i, colInd+l).value
            indAssignment.append(cellIndiv)
        assignmentList.append(indAssignment)
    return assignmentList


def getHeader(assignmentList):
    newHeader = []
    assignmentHeader = assignmentList.pop(0)
    newHeader.append(assignmentHeader[1])
    newHeader.append(assignmentHeader[0])
    newHeader.append('Days')
    return newHeader


def fullFormat(assignmentList, dayLimit):

    def fullFormatMain():
        onlyCompleted = removeDone(assignmentList)
        onlyCompleted.sort(key=listDays)
        twoWeek = twoWeekList(onlyCompleted, dayLimit)
        finalList = convertDate(twoWeek)
        return finalList

    def removeDone(assignmentList):
        onlyCompleted = [item for item in assignmentList if item[2] == None]
        return onlyCompleted

    def listDays(listToSort):
        return listToSort[0]

    def formatTime(dateVar):

        def formatTimeMain(dateVar):
            if type(dateVar) == datetime.datetime:
                tdMonth = dateVar.month
                tdDay = dateVar.day
                dateList = [tdMonth, tdDay]
                return padNum(dateList)
            else:
                return dateVar

        def padNum(var):
            for i in range(len(var)):
                timeVar = str(var[i])
                if len(timeVar) < 2:
                    timeVar = "0" + timeVar
                var[i] = timeVar
            return comboTime(var)

        def comboTime(var):
            timeStr = "/"
            dateStr = var[0] + timeStr + var[1]
            return dateStr
        return formatTimeMain(dateVar)

    def twoWeekList(onlyCompleted, dayLimit):
        todayDate = datetime.datetime.today().replace(
            hour=0, minute=0, second=0, microsecond=0)
        twoWeekerList = []
        for item in onlyCompleted:
            daysBetween = (item[0] - todayDate)
            daysBetween = daysBetween.days
            if daysBetween < dayLimit:
                indItemList = []
                if daysBetween < 2:
                    indItemList.append(str(item[1]).upper())
                    indItemList.append(item[0])
                    if daysBetween == -1:
                        indItemList.append('YESTERDAY!')
                    elif daysBetween == 0:
                        indItemList.append('TODAY!')
                    elif daysBetween == 1:
                        indItemList.append('Tomorrow!')
                    else:
                        indItemList.append(str(daysBetween) + ' Days Ago!')
                    twoWeekerList.append(indItemList)
                else:
                    indItemList.append(str(item[1]))
                    indItemList.append(item[0])
                    indItemList.append(str(daysBetween)+' Days')
                    twoWeekerList.append(indItemList)
        return twoWeekerList

    def convertDate(twoWeekList):
        for i in twoWeekList:
            newDate = formatTime(i[1])
            i[1] = newDate
        return twoWeekList

    return fullFormatMain()


def styleTable(df, headerCol):
    cStyle = tableStyle()
    bPx = cStyle['brWidth']
    bCo = cStyle['brColor']
    hFnt = cStyle['fontHead']
    hSFnt = cStyle['fontHsize']
    bFnt = cStyle['fontBody']
    bSFnt = cStyle['fontBsize']
    hCo = cStyle['HeadColor']
    hFntCo = cStyle['hFntColor']
    bFntCo = cStyle['bFntColor']
    rECo = cStyle['rowCoE']
    rOCo = cStyle['rowCoO']

    paddingHead = 'padding-top: %sem; padding-bottom: %sem;' % (0, 0)
    propsHead = 'font-weight:normal; background-color: #%s; font-family: %s; color: #%s; font-size: %sem;' % (
        hCo, hFnt, hFntCo, hSFnt)

    paddingBody = 'padding-top: %sem; padding-bottom: %sem;' % (0.3, 0.3)
    paddingBodyL = 'padding-left: %sem; padding-top: %sem; padding-bottom: %sem;' % (
        1, 0.3, 0.3)
    propsBodyE = 'font-weight:normal; background-color: #%s; font-family: %s; color: #%s; font-size: %sem;' % (
        rECo, bFnt, bFntCo, bSFnt)
    propsBodyO = 'font-weight:normal; background-color: #%s; font-family: %s; color: #%s; font-size: %sem;' % (
        rOCo, bFnt, bFntCo, bSFnt)

    bhBottom = 'border-bottom: %spx solid #%s;' % ((bPx-2), rECo)
    bTop = 'border-top: %spx solid #%s;' % (bPx, bCo)
    bRight = 'border-right: %spx solid #%s;' % (bPx, bCo)
    bBottom = 'border-bottom: %spx solid #%s;' % (bPx, bCo)
    bLeft = 'border-left: %spx solid #%s;' % (bPx, bCo)

    styleList = [
        # COLOR
        {"selector": "th.col_heading",
            "props": f'{propsHead}; {paddingHead}; {bTop}; {bhBottom};'},

        {"selector": "tbody tr:nth-child(even)",
            "props": f'{propsBodyE};'},

        {"selector": "tbody tr:nth-child(odd)",
            "props": f'{propsBodyO};'},
        # ALIGNMENT
        {"selector": "th.col0",
            "props": f"text-align: left;{bLeft}; {paddingBodyL};"},

        {"selector": "td.col0",
            "props": f"text-align: left; {paddingBodyL}; {bLeft};"},

        {"selector": "th.col1",
            "props": "text-align: center;"},

        {"selector": "td.col1",
            "props": f"text-align: center; {paddingBody}"},

        {"selector": "th.col2",
            "props": f"text-align: center; {bRight}"},

        {"selector": "td.col2",
            "props": f"text-align: right; {paddingBody}; {bRight}"},

        {"selector": "tbody tr:nth-last-child(1)",
            "props": f"text-align: right; {paddingBody}; {bBottom}"},
    ]

    lengthofToday = len(df[df[headerCol[2]].str.contains("!")])

    if lengthofToday > 0:
        for i in range(lengthofToday):
            tempToday = {}
            dict_keys = ['selector', 'props']
            tempToday[dict_keys[0]] = ""
            tempToday[dict_keys[1]
                      ] = 'font-weight: bold; color: #%s;' % (cStyle['pastCo'])
            tempToday['selector'] = (f"tbody tr:nth-child({i+1})")
            styleList.append(tempToday)
            del tempToday
    return styleList

# --------------------------------------------------------------#


if __name__ == '__main__':
    main()
