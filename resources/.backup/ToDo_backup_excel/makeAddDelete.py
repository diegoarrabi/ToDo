from config import excelPath, myLog, path_dict, notifications_on
from subprocess import run
import os
import openpyxl
import datetime


def main() -> None:

    project_directory = path_dict['Project']
    wb_path = excelPath(project_directory)
    sheet_name = "Assignments"
    txt_doc_name = "AssignmentEdit.txt"

    for item in txtFileData(txt_doc_name, project_directory):
        if item == "_skip":
            break
        workbook = openpyxl.load_workbook(wb_path)
        worksheet = workbook[sheet_name]
        assignment_data = item.split(" - ")
        duedate_value = (assignment_data[1].strip()).lower()
        myLog(f'Task: {assignment_data}')

        if duedate_value != 'done':
            assignmentAddEdit(worksheet, assignment_data)
        elif duedate_value == 'done':
            assignmentDone(worksheet, assignment_data)

        workbook.save(wb_path)
        workbook.close()

    py_script = "makeTable.py"
    script = os.path.join(project_directory, py_script)
    run(['python3', script])

    myLog('----------------------------DONE----------------------------\n\n')


def txtFileData(txtfile_name: str, basedir: str) -> list[str]:
    """
    Reads Assignment text file and returns a list of each line (stripped of whitespaces)

    Args:
        txtfile_name (str): name of text file
        txtfile_path (str): path of parent folder

    Returns:
        list[str]: each line as a value 
    """

    txtfile_path = os.path.join(basedir, txtfile_name)
    with open(txtfile_path, "r+") as read_file:
        txt_data = read_file.readlines()
        txt_data = [x.strip() for x in txt_data]
        read_file.truncate(0)
        if len(txt_data) == 0:
            txt_data = ["_skip"]
    return txt_data


def assignmentAddEdit(worksheet, assignment_info: list) -> None:

    def addDate(count: int, adddelete: str) -> None:
        cell_adddate = worksheet.cell(count, date_column)
        cell_adddate.value = datetime.datetime(item_date[2], item_date[0], item_date[1])
        cell_adddate.number_format = "MM/DD/YY"
        myLog(f'Cell Index: [ {count} , {date_column} ]')
        showNotification(assignment_info[0], adddelete)

    myLog("ADDING/EDITING ASSIGNMENT")

    first_row = 1
    date_column = 1
    assignment_column = 2
    item_date = getFullDate(assignment_info)

    iterate_rows = True
    iterate_count = 0

    while iterate_rows == True:

        row_iteration = first_row + iterate_count
        assignment_iteration = worksheet.cell((row_iteration), assignment_column).value

        # assignment_type = type(assignment_iteration)
        # myLog(f"Assignment {iterate_count}: {assignment_iteration}")
        # myLog(f"Assignment Type {assignment_type}")
        # myLog(f'Cell Index: [ {row_iteration} , {assignment_column} ]\n')

        if assignment_iteration is not None:
            assignment_iteration = assignment_iteration.lower()

        if assignment_iteration == assignment_info[0].lower():
            myLog("EDITING ASSIGNMENT")
            adddelete = 'edit'
            addDate(row_iteration, adddelete)
            myLog(f'Edit Task: {assignment_info[0]} -> {assignment_info[1]}\n')
            break

        elif assignment_iteration is None:
            myLog("ADDING ASSIGNMENT")
            adddelete = 'add'
            cell_index = worksheet.cell(row_iteration, assignment_column)
            cell_index.value = assignment_info[0]
            addDate(row_iteration, adddelete)
            myLog(f'Add Task: {assignment_info[0]}\n')
            break

        if iterate_count == 100:
            break
        iterate_count += 1


def assignmentDone(worksheet, assignment_info: list) -> int:

    myLog("REMOVING ASSIGNMENT\n")

    first_row = 1
    assignment_column = 2

    adddelete = 'delete'

    iterate_count = 0
    iterate_rows = True

    while iterate_rows == True:

        row_iteration = first_row + iterate_count
        assignment_iteration = worksheet.cell((row_iteration), (assignment_column)).value

        # assignment_type = type(assignment_iteration)
        # myLog(f"Assignment {iterate_count}: {assignment_iteration}")
        # myLog(f"Assignment Type {assignment_type}")
        # myLog(f'Cell Index: [ {row_iteration} , {assignment_column} ]\n')

        if assignment_iteration is not None:
            assignment_iteration = (assignment_iteration).lower()

        if assignment_iteration == assignment_info[0].lower():
            myLog(f'Delete Task: {assignment_info[0]}')
            myLog(f'Return Value [Row #]: {row_iteration}\n')
            showNotification(assignment_info[0], adddelete)
            worksheet.delete_rows(row_iteration)
            return (row_iteration)

        if iterate_count == 100:
            break
        iterate_count += 1


def getFullDate(item_info) -> list[int]:
    today_date = datetime.datetime.today()
    today_str = str(today_date.day)
    if len(today_str) < 2:
        today_str = '0' + today_str
    today_month = str(today_date.month)
    today_compare = int(today_month + today_str)

    item_index = str(item_info[1]).split('/')
    item_day = item_index[1]
    if len(item_day) < 2:
        item_day = '0' + item_day
    item_month = item_index[0]
    item_compare = int(item_month + item_day)

    if today_compare <= item_compare:
        item_year = today_date.year
    elif today_compare > item_compare:
        item_year = (today_date.year + 1)
    return [int(item_month), int(item_day), int(item_year)]


def showNotification(item_name, add_delete) -> None:
    if notifications_on:
        if add_delete == 'add':
            add_delete = "Added"
        elif add_delete == 'delete':
            add_delete = "Finished"
        elif add_delete == 'edit':
            add_delete = 'Changed Date'
        s1 = '-e set dText to "To Do List"'
        s2 = '-e set dTitle to "%s"' % (add_delete)
        s3 = '-e set dSub to "%s"' % (item_name)
        s4 = '-e set dSound to "Hero"'
        s5 = '-e Display Notification dText with title dTitle subtitle dSub sound name dSound'
        run(['osascript', s1, s2, s3, s4, s5])


if __name__ == '__main__':
    main()
