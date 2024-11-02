import os
import sys
import openpyxl

def errorOccured(errorVar, wbName="", basedir="", excelPath=""):
    if errorVar == 1:
        print("----File Not Found----")
        print("-----------------------")
        print(f"File: {wbName}")
        print(f"Directory: {basedir}")
        print("-----------------------")
        print(f"FullPath: {excelPath}")
        print("-----------------------")
    elif errorVar == 2:
        print(f"Sheet: '{wbName}' not found")
    sys.exit()
    return

def openExcelSheet(wbName, ssName):
    wbName = "/" + str(wbName)
    basedir = os.path.dirname(__file__)
    excelPath = os.path.join(basedir+wbName)
    try:
        workbook = openpyxl.load_workbook(excelPath, data_only=True)
    except FileNotFoundError:
        errorOccured(1, wbName, basedir, excelPath)
        return
    sheetInt = isofInterest(workbook, ssName)
    return sheetInt

def isofInterest(wb, ssName):
    for sheets in wb:
        sName = str(sheets)
        sName = sName.split('"', 2)
        if sName[1] == str(ssName):
            wbSheet = sheets
            return wbSheet
    errorOccured(2, ssName)

def findMatrixLU(wb):
    def checkColumn(wb, numVar):
        columnValue = True
        i = 0
        while columnValue != None:
            cellColumn = wb.cell(numVar, numVar-i).value
            if cellColumn == None:
                colIndex = (numVar-i+1)
                return colIndex
            i += 1
            
    def checkRow(wb, numVar):
        rowValue = True
        i = 0
        while rowValue != None:
            cellRow = wb.cell(numVar-i, numVar).value
            if cellRow == None:
                rowIndex = (numVar-i+1)
                return rowIndex
            i += 1

    def assigCount(wb, rowIndex, colIndex):
        rowStart = rowIndex
        rowValue = False
        i = 0
        while rowValue == False:
            cellRow = wb.cell(rowIndex+i, colIndex+3).value
            if cellRow == None:
                if wb.cell(rowIndex+i+1, colIndex+3).value != None:
                    pass
                elif wb.cell(rowIndex+i+1, colIndex+3).value == None:
                    return (rowIndex+i-rowStart)
            i += 1
        return

    hasValue = False
    i = 1
    rowI = 0
    while hasValue == False:
        cellMatrix = wb.cell(rowI + i,i)
        cellValue = cellMatrix.value
        if i > 20:
            rowI += 1
            i = 0
        if cellValue != None:
            colIndex = checkColumn(wb, cellMatrix.column) 
            rowIndex = checkRow(wb, cellMatrix.row)
            lastCell = assigCount(wb, rowIndex, colIndex)
            hasValue = True
        i += 1 
    return rowIndex, colIndex, lastCell

def makeAssg(wb, startTable):
    assignmentList = []
    rowInd = startTable[0]
    colInd = startTable[1]
    assLen = startTable[2]
    for i in range(assLen):
        indAssignment = []
        for l in range(5):
            cellIndiv = wb.cell(rowInd+i,colInd+l).value
            indAssignment.append(cellIndiv)
        assignmentList.append(indAssignment)
    return assignmentList

def getHeader(assignmentList):
    assignmentHeader = assignmentList.pop(0)
    return assignmentHeader

def fullFormat(assignmentList):

    def formatList(assignmentList):
        removeBlank = [item for item in assignmentList if item[0] != None]
        removeDone = [item for item in removeBlank if item[1] != None]
        return removeDone
    
    def listDays(listToSort):
        return listToSort[1]
    
    def formatTime(dateVar):
        def padNum(var):
            for i in range(len(var)):
                timeVar = str(var[i])
                if len(timeVar) < 2:
                    timeVar = "0" + timeVar
                var[i] = timeVar
            return comboTime(var)
        def comboTime(var):
            timeStr = "/"
            dateStr = var[0] + timeStr + var[1]
            return dateStr

        tdMonth = dateVar.month
        tdDay = dateVar.day
        dateList = [tdMonth, tdDay]
        return padNum(dateList)
    
    def make2Week():
        newMasterAssig = []
        for assignInfo in formatMasterList:    
            if assignInfo[1] < 14:
                newAssig = []
                newDate = formatTime(assignInfo[2])
                newAssig.append(assignInfo[3])
                newAssig.append(newDate)
                newAssig.append(assignInfo[1])
                newMasterAssig.append(newAssig)
        return newMasterAssig
    
    formatMasterList = formatList(assignmentList)
    formatMasterList.sort(key=listDays)
    finalTwoWeek = make2Week()

    return finalTwoWeek


workbookName = "pyToDo.xlsm"
sheetName = "Assignments"
wb = openExcelSheet(workbookName, sheetName)
startTable = findMatrixLU(wb)
assignmentList = makeAssg(wb, startTable)
assignmentHeader = getHeader(assignmentList)
twoWeekList = fullFormat(assignmentList)

##Pretty format left to do##

print(assignmentHeader)
print()

# for i in twoWeekList:
#     print(i)

for assignment in twoWeekList:
    if assignment[2] == 0:
        assignment[2] = "TODAY!"
        assignment[0] = assignment[0].upper()
    elif assignment[2] == 1:
        assignment[2] = "Tomorrow!"
        assignment[0] = assignment[0].upper()
    else:
        assignment[2] = str(assignment[2]) + " Days"

for i in twoWeekList:
    print(i)

# myFile = open('todocsv.csv', 'w')
# writer = csv.writer(myFile)
# for assignmentItem in twoWeekList:
#     writer.writerow(assignmentItem)
# myFile.close()
